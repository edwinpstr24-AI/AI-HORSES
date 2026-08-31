#!/usr/bin/env python3
"""
AI Horses - extract_selections.py

Reads the SELECTIONS sheet out of a scoring workbook and prints the
engine_output blocks ready to paste into a card JSON.

Usage:
    python extract_selections.py "GULFSTREAM MASTER NEW .xlsm"
    python extract_selections.py workbook.xlsm --races 1-8
    python extract_selections.py workbook.xlsm --merge CARDS/CT-2026-08-29-E.json

The workbook is opened read-only and never written to. Run it AFTER you have
opened the file in Excel and pressed the sort button, so the cached values are
current - openpyxl reads what Excel last computed, not a live recalculation.

SELECTIONS layout (confirmed against the Gulfstream master):
    14 race blocks, 7 rows apart, first block at row 24.
    Odd races on the left  (labels col B, data cols C-G)
    Even races on the right (labels col I, data cols J-N)
    block+2 = program numbers, ranks 1st/2nd/3rd/4th/longshot
    block+3 = scores
    block+4 = TOP CLASS (col D/K), TOP POWER (col G/N)
    block+5 = TOP PACE  (col D/K), TOP SPEED (col G/N)
"""

import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed.  pip install openpyxl")

SHEET = "SELECTIONS "
FIRST_ROW = 24
STEP = 7
LEFT = {"cols": [3, 4, 5, 6, 7], "cat_a": 4, "cat_b": 7}    # C-G
RIGHT = {"cols": [10, 11, 12, 13, 14], "cat_a": 11, "cat_b": 14}  # J-N


def cell(ws, row, col):
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def as_program(v):
    """Program numbers may come back as ints, floats or strings."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def block_row(race):
    return FIRST_ROW + ((race - 1) // 2) * STEP


def extract_race(ws, race):
    row = block_row(race)
    side = LEFT if race % 2 == 1 else RIGHT
    cols = side["cols"]

    ranked = []
    for i in range(4):
        prog = as_program(cell(ws, row + 2, cols[i]))
        if prog is None:
            continue
        entry = {"rank": i + 1, "program": prog}
        score = cell(ws, row + 3, cols[i])
        if isinstance(score, (int, float)):
            entry["score"] = round(float(score), 2)
        ranked.append(entry)

    out = {}
    if ranked:
        out["ranked"] = ranked

    ls_prog = as_program(cell(ws, row + 2, cols[4]))
    if ls_prog:
        ls = {"program": ls_prog}
        ls_score = cell(ws, row + 3, cols[4])
        if isinstance(ls_score, (int, float)):
            ls["score"] = round(float(ls_score), 2)
        out["longshot"] = ls

    for key, r, c in [
        ("top_class", row + 4, side["cat_a"]),
        ("top_power", row + 4, side["cat_b"]),
        ("top_pace", row + 5, side["cat_a"]),
        ("top_speed", row + 5, side["cat_b"]),
    ]:
        v = cell(ws, r, c)
        if v:
            out[key] = str(v)

    return out or None


def extract_day_picks(ws):
    """
    Row 22 pairs label cells with input cells you fill in by hand:
        D22 'RACE' -> E22 value      F22 'HORSE #' -> G22 value
        K22 'RACE' -> L22 value      M22 'HORSE #' -> N22 value
    """
    LABELS = {"RACE", "HORSE #", "BEST BET OF THE DAY", "LONGSHOT OF THE DAY"}

    def clean(v):
        p = as_program(v)
        if p is None or p.strip().upper() in LABELS:
            return None
        return p

    picks = {}
    for key, race_col, prog_col in [("best_bet", 5, 7), ("longshot_of_day", 12, 14)]:
        prog = clean(cell(ws, 22, prog_col))
        if not prog:
            continue
        entry = {"program": prog}
        race = cell(ws, 22, race_col)
        if isinstance(race, (int, float)):
            entry["race"] = int(race)
        picks[key] = entry
    return picks or None


def parse_races(spec):
    if not spec:
        return list(range(1, 15))
    out = []
    for part in spec.split(","):
        if "-" in part:
            a, b = part.split("-")
            out.extend(range(int(a), int(b) + 1))
        else:
            out.append(int(part))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--races", help="e.g. 1-8 or 1,3,5. Default all 14.")
    ap.add_argument("--merge", help="Card JSON to write engine_output into.")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        match = [n for n in wb.sheetnames if n.strip().upper() == "SELECTIONS"]
        if not match:
            sys.exit(f"No SELECTIONS sheet. Found: {wb.sheetnames}")
        ws = wb[match[0]]
    else:
        ws = wb[SHEET]

    races = parse_races(args.races)
    extracted = {}
    empty = []
    for n in races:
        block = extract_race(ws, n)
        if block and block.get("ranked"):
            extracted[n] = block
        else:
            empty.append(n)

    day_picks = extract_day_picks(ws)

    if empty:
        print(f"# No scored data for race(s): {', '.join(map(str, empty))}",
              file=sys.stderr)
    if not extracted:
        sys.exit("Nothing extracted. Open the workbook in Excel, run the sort, "
                 "save, then try again.")

    # All-zero scores mean the sheet was never computed.
    scores = [e.get("score", 0) for b in extracted.values() for e in b.get("ranked", [])]
    if scores and all(s == 0 for s in scores):
        print("# WARNING: every score is 0. The workbook looks uncomputed - "
              "open it in Excel, run the sort, save, and re-run.", file=sys.stderr)

    if args.merge:
        card = json.load(open(args.merge, encoding="utf-8"))
        hit = 0
        for race in card.get("races", []):
            n = race.get("race_number")
            if n in extracted:
                race["engine_output"] = extracted[n]
                hit += 1
        if day_picks:
            card["day_picks"] = day_picks
        json.dump(card, open(args.merge, "w", encoding="utf-8"), indent=2)
        open(args.merge, "a", encoding="utf-8").write("\n")
        print(f"Merged engine_output into {hit} race(s) of {args.merge}")
        if day_picks:
            print("Added day_picks.")
        print("Re-run validate.py before committing.")
        return

    payload = {"day_picks": day_picks} if day_picks else {}
    payload["races"] = [{"race_number": n, "engine_output": b}
                        for n, b in sorted(extracted.items())]
    print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
